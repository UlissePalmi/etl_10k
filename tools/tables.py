# ============================================================================
# EXTRACT TABLES FROM FINANCIAL STATEMENTS (Item 8) AND CONVERT TO EXCEL
# ============================================================================

from bs4 import BeautifulSoup
from openpyxl import Workbook
from pathlib import Path

# ============================================================================
# STEP 1: READ THE HTML FILE
# ============================================================================
# Path to the cleaned Item 8 (Financial Statements) file
item8_path = Path(r"C:\Users\upalmier\Documents\etl_10k\data\interim\items\0001562088\10-K\0001562088-25-000042\item8.txt")

# Read the entire file as text (encoding='utf-8' handles special characters)
content = item8_path.read_text(encoding='utf-8')

# ============================================================================
# STEP 2: PARSE HTML AND FIND ALL TABLES
# ============================================================================
# BeautifulSoup parses the HTML content
soup = BeautifulSoup(content, 'html.parser')

# Find ALL <table> tags in the document
tables = soup.find_all('table')

# Print how many tables we found
print(f"Found {len(tables)} tables in item8.txt")

# ============================================================================
# STEP 3: HELPER FUNCTION TO PARSE NUMERIC VALUES
# ============================================================================
def parse_val(s):
    """
    Convert cell text to a number if possible.
    Handles:
    - Removes commas: "1,234,567" → "1234567"
    - Removes spaces: "1 234 567" → "1234567"
    - Converts negatives: "(123)" → -123.0
    - Returns None if conversion fails
    """
    s = s.strip().replace(',', '').replace(' ', '')

    # Check if number is in parentheses (accounting format for negatives)
    if s.startswith('(') and s.endswith(')'):
        try:
            return -float(s[1:-1])  # Remove parens and negate
        except:
            return None

    # Try to convert to float
    try:
        return float(s)
    except:
        return None

# ============================================================================
# STEP 4: CREATE EXCEL WORKBOOK
# ============================================================================
# Create a new Excel workbook
wb = Workbook()

# Remove the default empty sheet that openpyxl creates
if 'Sheet' in wb.sheetnames:
    wb.remove(wb['Sheet'])

# ============================================================================
# STEP 5: PROCESS EACH TABLE
# ============================================================================
# Loop through each table and create a separate sheet for it
for table_idx, table in enumerate(tables, 1):  # Start counting from 1
    rows = []
    row_num = 0

    # For each row in the table
    for tr in table.find_all('tr'):
        row_num += 1

        # Extract text from each cell (<td>) in the row
        cells = [td.get_text(' ', strip=True).replace('\n', ' ').strip() for td in tr.find_all('td')]
        # Explanation:
        #   td.get_text(' ', strip=True) = get text from cell, join nested text with space
        #   strip=True = remove leading/trailing whitespace from cell
        #   .replace('\n', ' ') = replace newlines with spaces (for multi-line cells)
        #   .strip() = final cleanup of whitespace

        # Print detailed output for Table 2
        if table_idx == 2:
            print(f"\n--- TABLE 2, ROW {row_num} ---")
            print(f"Raw cells extracted: {cells}")
            print(f"Total cells: {len(cells)}")

        # Filter cells: remove empty strings and '$' symbols
        # NOTE: This causes column misalignment if rows have different numbers of cells!
        cells = [c for c in cells if c and c != '$']

        # Only add row if it has at least one non-empty cell
        if cells:
            rows.append(cells)
            if table_idx == 2:
                print(f"Filtered cells: {cells}")
                print(f"Status: ✓ ADDED (length: {len(cells)})")
        else:
            if table_idx == 2:
                print(f"Filtered cells: EMPTY")
                print(f"Status: ✗ SKIPPED (no non-empty cells)")

    # Create a new sheet in the workbook for this table
    ws = wb.create_sheet(title=f"Table {table_idx}")

    # Write all extracted rows to the sheet
    for row in rows:
        ws.append(row)  # Each item in 'row' goes into a separate column (A, B, C, etc.)

    # Print progress
    print(f"\nTable {table_idx}: {len(rows)} rows")

# ============================================================================
# STEP 6: SAVE WORKBOOK
# ============================================================================
# Save the workbook with all tables to a file
wb.save('duolingo_all_tables.xlsx')
print("Saved to duolingo_all_tables.xlsx")
