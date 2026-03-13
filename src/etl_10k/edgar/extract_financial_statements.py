"""
Extract all R*.htm financial statement tables from raw SGML 10-K filings.

For each filing, parses FilingSummary.xml to get the authoritative R-file → ShortName mapping,
then extracts the main <table class="report"> from each R-file and saves to Excel.

Output: data/interim/financial_statements/<CIK>/<accession>/financial_statements.xlsx
Each workbook has one sheet per R-file (up to 75 sheets for Duolingo).
"""

import re
import argparse
from pathlib import Path
from bs4 import BeautifulSoup
from openpyxl import Workbook

from etl_10k.config import RAW_EDGAR_DIR, INTERIM_DIR


def _extract_tag(text, tag):
    """Extract first <tag>...</tag> match from text. Returns '' if not found."""
    pattern = f"<{tag}>(.*?)</{tag}>"
    match = re.search(pattern, text, re.DOTALL)
    return match.group(1) if match else ""


def parse_val(s):
    """
    Convert cell text to a number if possible.
    Handles:
    - Removes commas: "1,234,567" → "1234567"
    - Removes spaces: "1 234 567" → "1234567"
    - Converts negatives: "(123)" → -123.0
    - Returns None if conversion fails
    """
    s = s.strip().replace(",", "").replace(" ", "")

    # Check if number is in parentheses (accounting format for negatives)
    if s.startswith("(") and s.endswith(")"):
        try:
            return -float(s[1:-1])
        except:
            return None

    # Try to convert to float
    try:
        return float(s)
    except:
        return None


def parse_filing_summary(content):
    """
    Find <FILENAME>FilingSummary.xml block in SGML content, extract all <Report> blocks,
    return list of (HtmlFileName, ShortName, MenuCategory) tuples in document order.
    Skip entries with no HtmlFileName (e.g. Book-type "All Reports" entry).
    Returns [] for pre-XBRL filings.
    """
    # Find the start of FilingSummary.xml document
    fs_start = content.find("<FILENAME>FilingSummary.xml")
    if fs_start == -1:
        return []

    # Find the <TEXT> tag after this
    text_start = content.find("<TEXT>", fs_start)
    if text_start == -1:
        return []

    # Extract content between <TEXT> and </TEXT>
    text_content_start = text_start + 6  # len("<TEXT>")
    text_end = content.find("</TEXT>", text_content_start)
    if text_end == -1:
        return []

    fs_xml = content[text_content_start:text_end]

    # Extract all <Report> blocks (may have attributes like instance="...")
    reports = []
    for report_block in re.finditer(r"<Report[^>]*>(.*?)</Report>", fs_xml, re.DOTALL):
        report_xml = report_block.group(1)

        # Extract fields
        html_filename = _extract_tag(report_xml, "HtmlFileName")
        short_name = _extract_tag(report_xml, "ShortName")
        menu_category = _extract_tag(report_xml, "MenuCategory")

        # Skip entries with no HtmlFileName (e.g. Book-type entries)
        if html_filename:
            reports.append((html_filename, short_name, menu_category))

    return reports


def extract_document_html(content, filename):
    """
    Extract the HTML content of a specific R-file from the SGML.
    Uses string-search (not regex) — finds <FILENAME>{filename} marker,
    then locates the next <TEXT>...</TEXT> block.
    Returns None if not found.
    """
    # Find the <FILENAME>marker
    marker = f"<FILENAME>{filename}"
    idx = content.find(marker)
    if idx == -1:
        return None

    # Find the next <TEXT> after the marker
    text_start = content.find("<TEXT>", idx)
    if text_start == -1:
        return None

    # Find the matching </TEXT>
    text_end = content.find("</TEXT>", text_start)
    if text_end == -1:
        return None

    # Extract HTML content (between <TEXT> and </TEXT>)
    html = content[text_start + 6 : text_end]  # +6 to skip "<TEXT>"
    return html


def extract_table_rows(html):
    """
    Extract table rows from XBRL viewer HTML.
    Uses soup.find('table', class_='report') — stable CSS class for main table.
    Extracts both <td> and <th> cells per row.
    Applies parse_val() to convert numeric strings to floats.
    Skips rows where all cells are empty.
    Returns list of lists (each inner list is a row).
    """
    soup = BeautifulSoup(html, "html.parser")
    table = soup.find("table", class_="report")

    if not table:
        return []

    rows = []
    for tr in table.find_all("tr"):
        # Extract text from both <td> and <th> cells
        cells = []
        for cell in tr.find_all(["td", "th"]):
            cell_text = cell.get_text(" ", strip=True).replace("\n", " ").strip()
            cells.append(cell_text)

        # Skip rows where all cells are empty
        if any(c.strip() for c in cells):
            rows.append(cells)

    return rows


def make_sheet_name(short_name, used_names):
    """
    Create a valid Excel sheet name from ShortName.
    - Truncate to 31 characters (Excel limit)
    - Strip invalid Excel sheet name chars: [ ] : * ? / \
    - Deduplicate: if truncated name collides, append _2, _3, etc.
    """
    # Strip invalid chars
    invalid_chars = r"[\[\]:*?/\\]"
    name = re.sub(invalid_chars, "", short_name)

    # Truncate to 31 chars
    name = name[:31]

    # Deduplicate
    final_name = name
    counter = 2
    while final_name in used_names:
        final_name = f"{name[:27]}_{counter}"  # Leave room for _N suffix
        counter += 1

    return final_name


def process_filing(sgml_path, output_path):
    """
    End-to-end extraction for one filing:
    1. Skip if file is empty
    2. Parse FilingSummary → skip if pre-XBRL
    3. For each R-file: extract HTML → parse table → write sheet
    4. Save workbook; return count of sheets written
    """
    # Skip if empty
    if sgml_path.stat().st_size == 0:
        return {"sheets": 0, "status": "empty"}

    # Read SGML content (try UTF-8, fallback to latin-1)
    try:
        content = sgml_path.read_text(encoding="utf-8")
    except UnicodeDecodeError:
        try:
            content = sgml_path.read_text(encoding="latin-1")
        except Exception as e:
            return {"sheets": 0, "status": f"read_error: {e}"}

    # Parse FilingSummary
    reports = parse_filing_summary(content)
    if not reports:
        return {"sheets": 0, "status": "pre_xbrl"}

    # Create output directory
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # Create workbook
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # Track used sheet names to avoid duplicates
    used_names = set()
    sheets_written = 0

    # Process each R-file
    for html_filename, short_name, menu_category in reports:
        # Skip Parenthetical tables
        if "Parenthetical" in short_name:
            continue

        # Extract HTML
        html = extract_document_html(content, html_filename)
        if not html:
            continue

        # Extract table rows
        rows = extract_table_rows(html)
        if not rows:
            continue

        # Create sheet with sanitized name
        sheet_name = make_sheet_name(short_name, used_names)
        used_names.add(sheet_name)

        ws = wb.create_sheet(title=sheet_name)

        # Write rows to sheet
        for row in rows:
            # Convert numeric strings to floats where possible
            converted_row = [parse_val(cell) if parse_val(cell) is not None else cell for cell in row]
            ws.append(converted_row)

        sheets_written += 1

    # Save workbook
    if sheets_written > 0:
        wb.save(output_path)
        return {"sheets": sheets_written, "status": "ok"}
    else:
        return {"sheets": 0, "status": "no_tables"}


def extract_and_save(sgml_content, output_path):
    """
    Extract all R*.htm financial statement tables from raw SGML content and save to Excel.

    Args:
        sgml_content (str): Raw SGML file content (as string, e.g., from file_content.read_text())
        output_path (Path): Output Excel file path

    Returns:
        dict: {"sheets": N, "status": "ok"|"pre_xbrl"|"no_tables"|"error"}
    """
    try:
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        return process_filing_from_string(sgml_content, output_path)
    except Exception as e:
        return {"sheets": 0, "status": f"error: {e}"}


def process_filing_from_string(sgml_content, output_path):
    """
    Process a filing from SGML content string (not from disk).
    Same logic as process_filing() but takes content string instead of file path.
    """
    # Parse FilingSummary
    reports = parse_filing_summary(sgml_content)
    if not reports:
        return {"sheets": 0, "status": "pre_xbrl"}

    # Create output directory
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # Create workbook
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # Track used sheet names to avoid duplicates
    used_names = set()
    sheets_written = 0

    # Process each R-file
    for html_filename, short_name, menu_category in reports:
        # Skip Parenthetical tables
        if "Parenthetical" in short_name:
            continue

        # Extract HTML
        html = extract_document_html(sgml_content, html_filename)
        if not html:
            continue

        # Extract table rows
        rows = extract_table_rows(html)
        if not rows:
            continue

        # Create sheet with sanitized name
        sheet_name = make_sheet_name(short_name, used_names)
        used_names.add(sheet_name)

        ws = wb.create_sheet(title=sheet_name)

        # Write rows to sheet
        for row in rows:
            # Convert numeric strings to floats where possible
            converted_row = [parse_val(cell) if parse_val(cell) is not None else cell for cell in row]
            ws.append(converted_row)

        sheets_written += 1

    # Save workbook
    if sheets_written > 0:
        wb.save(output_path)
        return {"sheets": sheets_written, "status": "ok"}
    else:
        return {"sheets": 0, "status": "no_tables"}


def main():
    parser = argparse.ArgumentParser(
        description="Extract all R*.htm tables from raw SGML 10-K filings to Excel"
    )
    parser.add_argument(
        "--cik", type=str, help="Single CIK (10-digit, e.g. 0001562088)"
    )
    parser.add_argument(
        "--ciks", type=str, help="Comma-separated CIKs (e.g. 0001562088,0000320193)"
    )
    args = parser.parse_args()

    # Determine which CIKs to process
    if args.cik:
        cik_list = [args.cik]
    elif args.ciks:
        cik_list = [c.strip() for c in args.ciks.split(",")]
    else:
        # Default: all CIKs in RAW_EDGAR_DIR
        cik_list = [d.name for d in RAW_EDGAR_DIR.iterdir() if d.is_dir()]
        cik_list.sort()

    # Process each CIK
    filing_count = 0
    sheet_total = 0
    status_counts = {}

    for cik in cik_list:
        cik_dir = RAW_EDGAR_DIR / cik / "10-K"
        if not cik_dir.exists():
            continue

        # Process each accession
        for accession_dir in sorted(cik_dir.iterdir()):
            if not accession_dir.is_dir():
                continue

            sgml_path = accession_dir / "full-submission.txt"
            if not sgml_path.exists():
                continue

            # Extract
            output_dir = INTERIM_DIR / "financial_statements" / cik / accession_dir.name
            output_path = output_dir / "financial_statements.xlsx"

            result = process_filing(sgml_path, output_path)

            filing_count += 1
            sheet_total += result["sheets"]
            status_counts[result["status"]] = status_counts.get(result["status"], 0) + 1

            # Print progress
            print(
                f"[{filing_count}] {cik}/{accession_dir.name} -> "
                f"sheets={result['sheets']} | {result['status']}"
            )

    # Print summary
    print(f"\n--- Summary ---")
    print(f"Total filings processed: {filing_count}")
    print(f"Total sheets extracted: {sheet_total}")
    for status, count in sorted(status_counts.items()):
        print(f"  {status}: {count}")


def extract_year_from_accession(accession: str) -> int:
    """
    Extract year from accession number format: CIK-YY-SEQUENCE.
    E.g., '0001562088-25-000042' -> 2025
    """
    parts = accession.split("-")
    if len(parts) >= 2:
        yy = parts[1]
        # Convert YY to YYYY: 00-99
        # 00-30 -> 2000-2030, 31-99 -> 1931-1999
        yy_int = int(yy)
        year = 2000 + yy_int if yy_int <= 30 else 1900 + yy_int
        return year
    return None


def consolidate_statements_by_type(ciks):
    """
    Consolidate per-filing financial statements by statement type.

    For each CIK, groups all sheets with the same name across multiple 10-K filings
    into separate Excel workbooks (one per statement type). Sheet names in consolidated
    files are the fiscal years extracted from accession numbers.

    Structure:
        BEFORE: financial_statements/<CIK>/<accession>/financial_statements.xlsx
                  (one file per filing, multiple sheets per file)
        AFTER:  financial_statements/<CIK>/<statement_name>.xlsx
                  (one file per statement type, multiple years per file)

    Args:
        ciks: List of CIK strings to consolidate
    """
    from openpyxl import load_workbook
    import shutil

    cik_list = list(ciks)
    financial_statements_root = INTERIM_DIR / "financial_statements"

    if not financial_statements_root.exists():
        print("No financial statements extracted yet.")
        return

    total_consolidated = 0

    for cik in cik_list:
        padded_cik = str(cik).zfill(10)
        cik_dir = financial_statements_root / padded_cik
        if not cik_dir.exists():
            print(f"  No statements for {cik}")
            continue

        # Collect all sheets from all accessions for this CIK
        # Group by base name (strip trailing _2, _3 dedup suffixes added by openpyxl)
        # Format: {base_name: {(year, suffix): [[cell_values]]}}
        sheets_by_base = {}
        accession_dirs = []

        for accession_dir in sorted(cik_dir.iterdir()):
            if not accession_dir.is_dir():
                continue

            filing_xlsx = accession_dir / "financial_statements.xlsx"
            if not filing_xlsx.exists():
                print(f"    No Excel in {accession_dir.name}")
                continue

            accession_dirs.append(accession_dir)

            # Extract year from accession
            year = extract_year_from_accession(accession_dir.name)
            if year is None:
                continue

            # Read all sheets from this filing's Excel
            try:
                wb = load_workbook(filing_xlsx, data_only=True)
                for sheet_name in wb.sheetnames:
                    # Strip openpyxl dedup suffix (_2, _3, ...) to get base name
                    m = re.match(r'^(.*?)(_\d+)$', sheet_name)
                    base_name = m.group(1) if m else sheet_name
                    suffix = m.group(2) if m else ""  # e.g. "_2", "_3", or ""

                    if base_name not in sheets_by_base:
                        sheets_by_base[base_name] = {}

                    # Extract actual cell data (not references)
                    ws = wb[sheet_name]
                    sheet_data = []
                    for row in ws.iter_rows(values_only=True):
                        sheet_data.append(list(row))

                    sheets_by_base[base_name][(year, suffix)] = sheet_data

                wb.close()
            except Exception as e:
                print(f"  Error reading {filing_xlsx}: {e}")
                continue

        # Create consolidated Excel files, one per base statement name
        for base_name, year_suffix_data in sorted(sheets_by_base.items()):
            output_xlsx = cik_dir / f"{base_name}.xlsx"

            # Create new workbook for this statement type
            consolidated_wb = Workbook()
            consolidated_wb.remove(consolidated_wb.active)  # Remove default sheet

            # Add a sheet for each (year, suffix) in sorted order
            for (year, suffix) in sorted(year_suffix_data.keys()):
                sheet_data = year_suffix_data[(year, suffix)]
                sheet_title = f"{year}{suffix}"  # e.g. "2024", "2024_2", "2024_3"
                dest_ws = consolidated_wb.create_sheet(title=sheet_title)

                # Write data to destination sheet
                for row_idx, row_data in enumerate(sheet_data, start=1):
                    for col_idx, cell_value in enumerate(row_data, start=1):
                        dest_ws.cell(row=row_idx, column=col_idx, value=cell_value)

            # Save consolidated file
            consolidated_wb.save(output_xlsx)
            total_consolidated += 1
            years = sorted(set(y for y, _ in year_suffix_data.keys()))
            print(f"  {padded_cik}/{base_name}.xlsx ({len(years)} years)")

        # Delete per-filing directories (cleanup)
        for accession_dir in accession_dirs:
            try:
                shutil.rmtree(accession_dir)
            except Exception as e:
                print(f"  Warning: Could not delete {accession_dir}: {e}")

    print(f"\nConsolidated {total_consolidated} statement types")


if __name__ == "__main__":
    main()
